""" Capture the same location within multiple spreadsheets for easy cross-model comparison.
Useful for determining if you are looking at a one-off issue, or for determining if some
old code is still needed.
Output is a spreadsheet containing one tab, with the requested section repeated for each
input file.  The section is repeated twice: first the data (which is useful to make sure
you've 'registered' the same location properly, and second with the formlas *as strings*.
"""
from pathlib import Path
import openpyxl
import sys
import warnings

# openpyxl conmplains every time it encounters any widget it doesn't understand, and we just don't care.
if not sys.warnoptions:
    warnings.simplefilter("ignore")


def copy_xls_region(xlsfile, sheet, region, resultsheet, startrow, startcol, data_mode, format_mode):
    """scan the xls file in the requested mode, appending the requested region to resultsheet, 
    starting at startcol.
    data_mode: boolean for data: True copies the data in the cells, false copies the formulas
    format_mode: if True, copy cell formatting as well"""

    wb = openpyxl.load_workbook(xlsfile, read_only=True, data_only=data_mode, keep_links=False)
    if sheet not in wb.sheetnames:
        resultsheet.cell(startrow, startcol).value = f"Sheet {sheet} not found"
        return

    wbsheet = wb[sheet]
    cells = wbsheet[region]
    # cells is a list-of-lists of Cell or MergedCell objects

    torow = startrow
    for cellrow in cells:
        torow = torow + 1
        tocol = startcol
        for cell in cellrow:
            resultcell = resultsheet.cell(torow, tocol)
            resultcell.value = cell.value if data_mode else copy_formula(cell.value)
            if format_mode:
                copy_format(cell, resultcell)
            tocol = tocol + 1


def copy_format(from_cell, to_cell):
    to_cell.alignment = from_cell.alignment
    to_cell.border = from_cell.border
    to_cell.fill = from_cell.fill
    to_cell.number_format = from_cell.number_format

def copy_formula(value):
    """If the value is a formula, put a single quote in front of it to turn it into a string"""
    if (str(value).startswith('=')):
        return "'" + str(value)
    return str(value)


def sample_regions(filelist, sheetname, region, copy_data=True, copy_formula=True):
    """For each file in the list, copy the specified sheet&region to a new Workbook.
    Returns the workbook.  
    Copies the data, formulas (as strings), or both."""

    wb = openpyxl.Workbook()
    wbsheet = wb[wb.sheetnames[0]] # get the first sheet
    wbsheet.title = sheetname + " sample"

    (min_col, min_row, max_col, max_row) = openpyxl.utils.cell.range_boundaries(region)
    region_width = (max_col - min_col) + 1
    region_height = (max_row - min_row) + 1

    startcol = 2
    if copy_data:
        print("DATA")
        startrow = 2
        for file in filelist:
            fname = Path(file).name
            print(fname)
            # output file name
            wbsheet.cell(startrow, 2).value = fname
            copy_xls_region(file, sheetname, region, wbsheet, startrow+1, startcol, data_mode=True, format_mode=True)
            startrow = startrow + region_height + 4
        
        startcol = startcol + region_width + 4
    
    if copy_formula:
        print("FORMULAS")
        startrow = 2
        for file in filelist:
            fname = Path(file).name
            print(fname)
            if startcol == 2:  # there was no data column, so output file name now.
                wbsheet.cell(startrow, 2).value = fname
            copy_xls_region(file, sheetname, region, wbsheet, startrow+1, startcol, data_mode=False, format_mode=False)
            startrow = startrow + region_height + 4

    return wb



if __name__ == "__main__":
    import argparse
    import glob

    parser = argparse.ArgumentParser(
        description='Sample the same region across multiple Excel workbooks')
    parser.add_argument('--sheet', help='Excel sheet to look at')
    parser.add_argument('--region', help='Region inside sheet to look at in Excel notation (e.g. B14:D15)')
    parser.add_argument('--out', required=False, default="sample.xls", help='Where to put the output Excel; defaults to sample.out')
    parser.add_argument('files', nargs='+', help='Excel files to search')
    args = parser.parse_args()

    # Windows won't expand all glob'd paths, so do it ourselves.
    files = []
    for f in args.files:
        files.extend( glob.glob(f) )

    wb = sample_regions(files, args.sheet, args.region)
    wb.save(args.out)